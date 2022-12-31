import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_wb(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        val = sheet.cell(row, 3)
        new_price = val.value * 0.9
        new_price_cell = sheet.cell(row, 4)
        new_price_cell.value = new_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    newname = filename[0:-5] + '_new.xlsx'
    wb.save(newname)


process_wb('transactions.xlsx')
